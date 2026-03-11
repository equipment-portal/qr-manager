import streamlit as st
import pandas as pd
import qrcode
import os
import urllib.request
import urllib.parse
from pathlib import Path
from datetime import datetime
import io
import base64
import json
import shutil

# --- Excel操作用ライブラリ ---
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# --- 画像処理用ライブラリ ---
from PIL import Image, ImageDraw, ImageFont, ImageOps

# --- 初期設定 ---
DB_CSV = Path("devices.csv")
QR_DIR = Path("qr_codes")
MANUAL_DIR = Path("manuals")
EXCEL_LABEL_PATH = Path("print_labels.xlsx")
LABEL_HISTORY_FILE = Path("label_history.json")
TEMP_LABEL_DIR = Path("temp_labels")

# 【追加】下書き画像を一時的に展開するフォルダ
DRAFT_IMG_DIR = Path("draft_images")

for d in [QR_DIR, MANUAL_DIR, TEMP_LABEL_DIR, DRAFT_IMG_DIR]:
    d.mkdir(exist_ok=True)

cloud_font_path = "BIZUDGothic-Regular.ttf"

def setup_fonts():
    if not os.path.exists(cloud_font_path):
        try:
            font_url = "https://github.com/googlefonts/morisawa-biz-ud-gothic/raw/main/fonts/ttf/BIZUDGothic-Regular.ttf"
            urllib.request.urlretrieve(font_url, cloud_font_path)
        except Exception as e:
            pass

setup_fonts()

def safe_filename(name):
    keepcharacters = (' ', '.', '_', '-')
    return "".join(c for c in name if c.isalnum() or c in keepcharacters).rstrip()

# ==========================================
# --- 画像自動圧縮＆最適化エンジン（スマホの回転防止もここで処理） ---
# ==========================================
def compress_image(uploaded_file, max_size=1000):
    try:
        if hasattr(uploaded_file, 'read'):
            file_bytes = uploaded_file.read()
            img = Image.open(io.BytesIO(file_bytes))
            uploaded_file.seek(0)
        else:
            img = Image.open(uploaded_file)
            
        img = ImageOps.exif_transpose(img) # ここでスマホ写真の90度回転バグを修正！
        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')
            
        img.thumbnail((max_size, max_size), Image.Resampling.LANCZOS)
        output = io.BytesIO()
        img.save(output, format="JPEG", quality=75, optimize=True)
        return output.getvalue()
    except Exception as e:
        print(f"圧縮エラー: {e}")
        return None

# ==========================================
# --- URL短縮 ＆ 爆速QR生成 ---
# ==========================================
def make_short_url(long_url):
    try:
        api_url = f"https://is.gd/create.php?format=simple&url={urllib.parse.quote(long_url)}"
        req = urllib.request.Request(api_url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req) as res:
            return res.read().decode('utf-8')
    except:
        return long_url

def make_optimized_qr(url):
    short_url = make_short_url(url)
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=1)
    qr.add_data(short_url)
    qr.make(fit=True)
    return qr.make_image(fill_color="black", back_color="white")

# ==========================================
# --- マニュアル画像 生成関数 ---
# ==========================================
def create_manual_image(data, output_path):
    W = 1600; margin = 80; content_w = W - margin * 2
    try:
        font_title = ImageFont.truetype(cloud_font_path, 80)
        font_sub = ImageFont.truetype(cloud_font_path, 55)
        font_text = ImageFont.truetype(cloud_font_path, 45)
    except:
        font_title = font_sub = font_text = ImageFont.load_default()

    sections = []
    header_img = Image.new('RGB', (W, 380), 'white')
    draw = ImageDraw.Draw(header_img)
    draw.rectangle([0, 0, W, 100], fill=(255, 215, 0))
    draw.text((W - margin, 25), f"管理番号: {data['id']}", fill="black", font=font_text, anchor="ra")
    draw.text((margin, 150), data['name'], fill="black", font=font_title)
    draw.rectangle([margin, 280, W - margin, 340], fill=(242, 155, 33))
    draw.text((margin + 20, 285), f"■ 使用電源: AC {data['power'] or '未設定'}", fill="white", font=font_text)
    sections.append(header_img)

    def process_img_section(img_file, title):
        if not img_file:
            sec_img = Image.new('RGB', (W, 200), 'white')
            s_draw = ImageDraw.Draw(sec_img)
            s_draw.text((margin, 20), title, fill="black", font=font_sub)
            s_draw.rectangle([margin, 90, W - margin, 190], outline="gray", width=3)
            s_draw.text((W // 2, 145), "画像なし", fill="gray", font=font_text, anchor="mm")
            return sec_img
        
        try:
            if isinstance(img_file, str):
                if img_file.startswith("http"):
                    req = urllib.request.Request(img_file, headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req) as res:
                        pil_img = Image.open(io.BytesIO(res.read()))
                else:
                    pil_img = Image.open(img_file)
            elif hasattr(img_file, 'read'):
                file_bytes = img_file.read()
                pil_img = Image.open(io.BytesIO(file_bytes))
                img_file.seek(0)
            else:
                pil_img = Image.open(img_file)
            
            pil_img = ImageOps.exif_transpose(pil_img).convert('RGB')
            new_h = int(content_w * (pil_img.height / pil_img.width))
            pil_img = pil_img.resize((content_w, new_h), Image.Resampling.LANCZOS)
            
            sec_img = Image.new('RGB', (W, 90 + new_h + 50), 'white')
            s_draw = ImageDraw.Draw(sec_img)
            s_draw.text((margin, 20), title, fill="black", font=font_sub)
            sec_img.paste(pil_img, (margin, 90))
            s_draw.rectangle([margin, 90, margin + content_w, 90 + new_h], outline="gray", width=3)
            return sec_img
        except: return None

    loto_suffix = "（関連機器、付帯設備）" if data.get('is_related_loto') else ""
    img_list = [
        (data.get('img_exterior'), "機器外観"),
        (data.get('img_outlet'), "コンセント位置"),
        (data.get('img_label'), "資産管理ラベル"),
        (data.get('img_loto1'), f"LOTO手順書{loto_suffix} Page 1"),
        (data.get('img_loto2'), f"LOTO手順書{loto_suffix} Page 2")
    ]
    for f, t in img_list:
        sec = process_img_section(f, t)
        if sec: sections.append(sec)

    total_h = sum(s.height for s in sections) + 100
    final_img = Image.new('RGB', (W, total_h), 'white')
    curr_y = 0
    for s in sections:
        final_img.paste(s, (0, curr_y))
        curr_y += s.height
    final_img.convert('RGB').save(output_path, format="JPEG", quality=85)

def create_manual_image_extended(data, extra_images, output_path):
    W = 1600; margin = 80; content_w = W - margin * 2
    try:
        font_sub = ImageFont.truetype(cloud_font_path, 65)
        font_text = ImageFont.truetype(cloud_font_path, 55)
    except: font_sub = font_text = ImageFont.load_default()

    create_manual_image(data, output_path)
    base = Image.open(output_path)
    added = []

    for ex_f, ex_t in extra_images:
        try:
            if isinstance(ex_f, str):
                if ex_f.startswith("http"):
                    req = urllib.request.Request(ex_f, headers={'User-Agent': 'Mozilla/5.0'})
                    with urllib.request.urlopen(req) as res: pil = Image.open(io.BytesIO(res.read()))
                else: pil = Image.open(ex_f)
            elif hasattr(ex_f, 'read'):
                file_bytes = ex_f.read()
                pil = Image.open(io.BytesIO(file_bytes))
                ex_f.seek(0)
            else: pil = Image.open(ex_f)

            pil = ImageOps.exif_transpose(pil).convert('RGB')
            nh = int(content_w * (pil.height / pil.width))
            pil = pil.resize((content_w, nh), Image.Resampling.LANCZOS)
            si = Image.new('RGB', (W, 160 + nh), 'white')
            dr = ImageDraw.Draw(si)
            dr.text((margin, 25), ex_t, fill="black", font=font_sub)
            si.paste(pil, (margin, 100))
            dr.rectangle([margin, 100, margin+content_w, 100+nh], outline="gray", width=3)
            added.append(si)
        except: continue

    memo_val = data.get("memo", "なし")
    dummy_draw = ImageDraw.Draw(Image.new('RGB', (1, 1)))
    lines = []
    line = ""
    for paragraph in memo_val.split('\n'):
        for char in paragraph:
            if dummy_draw.textbbox((0, 0), line + char, font=font_text)[2] <= (content_w - 60):
                line += char
            else:
                lines.append(line)
                line = char
        if line: lines.append(line); line = ""
    if not lines: lines = ["なし"]

    char_h = font_text.getbbox("あ")[3] - font_text.getbbox("あ")[1] if hasattr(font_text, 'getbbox') else font_text.getsize("あ")[1]
    line_step = char_h + 25
    memo_box_h = 110 + (len(lines) * line_step) + 60
    ms = Image.new('RGB', (W, memo_box_h), 'white')
    md = ImageDraw.Draw(ms)
    md.text((margin, 30), "■ メモ・備考", fill="black", font=font_sub)
    md.rectangle([margin, 110, W - margin, memo_box_h - 20], outline=(242, 155, 33), width=6)
    for i, l in enumerate(lines): md.text((margin + 40, 140 + (i * line_step)), l, fill="black", font=font_text)
    added.append(ms)

    final = Image.new('RGB', (W, base.height + sum(s.height for s in added) + 100), 'white')
    final.paste(base, (0, 0))
    cy = base.height
    for s in added: final.paste(s, (0, cy)); cy += s.height
    final.convert('RGB').save(output_path, format="JPEG", quality=85)


# ==========================================
# --- 印刷用ラベル ＆ Excel台帳 処理 ---
# ==========================================
def create_label_image(data):
    scale = 4; target_w_px = 350 * scale; target_h_px = 200 * scale
    try:
        font_title = ImageFont.truetype(cloud_font_path, 19 * scale) 
        font_main = ImageFont.truetype(cloud_font_path, 30 * scale)  
        font_sm = ImageFont.truetype(cloud_font_path, 12 * scale)    
        font_footer = ImageFont.truetype(cloud_font_path, 13 * scale) 
    except: font_title = font_main = font_sm = font_footer = ImageFont.load_default()
        
    device_name = data.get('name', '不明')
    device_power = data.get('power', '不明')
    label_img = Image.new('RGB', (target_w_px, target_h_px), 'white')
    draw = ImageDraw.Draw(label_img)
    draw.rectangle([0, 0, target_w_px - 1, target_h_px - 1], outline=(255, 255, 0), width=12 * scale)
    
    draw.text((18 * scale, 16 * scale), "■", fill="black", font=font_title)
    draw.text((42 * scale, 16 * scale), "機器情報・LOTO確認ラベル", fill="black", font=font_title)
    
    qr_size = 72 * scale
    if 'img_qr' in data and data['img_qr'] is not None:
        try:
            qr_pil = data['img_qr'].convert('RGB').resize((qr_size, qr_size))
            label_img.paste(qr_pil, (target_w_px - qr_size - 22 * scale, target_h_px - qr_size - 30 * scale))
        except: pass
    
    current_size = 30 * scale
    temp_font = font_main
    while (draw.textbbox((0, 0), device_name, font=temp_font)[2]) > (target_w_px - 40 * scale) and current_size > 12 * scale:
        current_size -= 1 * scale
        temp_font = ImageFont.truetype(cloud_font_path, current_size)

    draw.text((18 * scale, 52 * scale), "機器名称:", fill="black", font=font_sm)
    draw.text((18 * scale, 66 * scale), device_name, fill="black", font=temp_font)
    draw.text((18 * scale, 108 * scale), "使用電源:", fill="black", font=font_sm)
    draw.text((18 * scale, 122 * scale), f"AC {device_power}", fill="black", font=temp_font)
    draw.text((18 * scale, 172 * scale), "[QR] 詳細スキャン（外観・コンセント位置・LOTO手順）", fill="black", font=font_footer)
    
    return label_img.resize((350, 200), Image.Resampling.LANCZOS)

def rebuild_excel():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "印刷用ラベルシート"
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = ws.page_margins.right = ws.page_margins.top = ws.page_margins.bottom = 0.2
    
    history = []
    if LABEL_HISTORY_FILE.exists():
        try:
            with open(LABEL_HISTORY_FILE, "r", encoding="utf-8") as f: history = json.load(f)
        except: pass

    for count, item in enumerate(history):
        img_path = TEMP_LABEL_DIR / item["img_filename"]
        if not img_path.exists(): continue
        c_idx = count // 13; r_idx = count % 13  
        cell_col = c_idx + 1; cell_row = r_idx + 1
        col_letter = get_column_letter(cell_col)
        ws.column_dimensions[col_letter].width = 19.5
        ws.row_dimensions[cell_row].height = 63.0 
        xl_img = XLImage(str(img_path))
        xl_img.width = 132; xl_img.height = 76
        xl_img.anchor = f"{col_letter}{cell_row}"
        ws.add_image(xl_img)
    wb.save(EXCEL_LABEL_PATH)

def add_label_to_history(name, label_img):
    history = []
    if LABEL_HISTORY_FILE.exists():
        try:
            with open(LABEL_HISTORY_FILE, "r", encoding="utf-8") as f: history = json.load(f)
        except: pass
    fname = f"label_{datetime.now().strftime('%Y%m%d%H%M%S%f')}.png"
    label_img.save(TEMP_LABEL_DIR / fname, format='PNG')
    history.append({"name": name, "img_filename": fname})
    with open(LABEL_HISTORY_FILE, "w", encoding="utf-8") as f: json.dump(history, f, ensure_ascii=False, indent=2)
    rebuild_excel()

def delete_label_from_history(index):
    history = []
    if LABEL_HISTORY_FILE.exists():
        try:
            with open(LABEL_HISTORY_FILE, "r", encoding="utf-8") as f: history = json.load(f)
        except: pass
    if 0 <= index < len(history):
        try: (TEMP_LABEL_DIR / history[index]["img_filename"]).unlink()
        except: pass
        history.pop(index)
        with open(LABEL_HISTORY_FILE, "w", encoding="utf-8") as f: json.dump(history, f, ensure_ascii=False, indent=2)
        rebuild_excel()

def clear_history():
    try: EXCEL_LABEL_PATH.unlink()
    except: pass
    try: LABEL_HISTORY_FILE.unlink()
    except: pass
    for f in TEMP_LABEL_DIR.glob("*.png"):
        try: f.unlink()
        except: pass

# ==========================================
# --- ストレージ保存処理（個別画像用） ---
# ==========================================
def save_image_to_storage(file_obj, did, suffix, mode, repo, token, local_path):
    if not file_obj: return ""
    comp_data = compress_image(file_obj)
    if not comp_data: return ""
    
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    fname = f"{safe_filename(did)}_{suffix}_{timestamp}.jpg"
    
    if mode == "2. 全自動（データベース保存）":
        encoded = base64.b64encode(comp_data).decode("utf-8")
        api_url = f"https://api.github.com/repos/{repo}/contents/images/{fname}"
        payload = {"message": f"Upload {fname}", "content": encoded, "branch": "main"}
        req = urllib.request.Request(api_url, data=json.dumps(payload).encode("utf-8"), method="PUT")
        req.add_header("Authorization", f"token {token}")
        req.add_header("Content-Type", "application/json")
        try:
            with urllib.request.urlopen(req) as res:
                res_data = json.loads(res.read().decode("utf-8"))
                raw_url = res_data["content"]["html_url"]
                return raw_url.replace("https://github.com/", "https://cdn.jsdelivr.net/gh/").replace("/blob/", "@")
        except: return ""
        
    elif mode == "3. 社内共有フォルダへ自動保存":
        base_dir = Path(local_path) / "images"
        base_dir.mkdir(parents=True, exist_ok=True)
        out_path = base_dir / fname
        with open(out_path, "wb") as f: f.write(comp_data)
        return str(out_path).replace("\\", "/")
    
    return ""

# ==========================================
# --- マスター台帳Excelの自動生成・保存 ---
# ==========================================
def update_master_ledger_excel(df_csv, mode, repo, token, local_path):
    try:
        df_export = df_csv.rename(columns={
            "ID": "管理番号", "Name": "機器名称", "Power": "使用電源",
            "URL": "マニュアルURL", "Updated": "最終更新日時", "memo": "メモ・備考"
        })
        cols_to_keep = ["管理番号", "機器名称", "使用電源", "マニュアルURL", "最終更新日時", "メモ・備考"]
        df_export = df_export[[c for c in cols_to_keep if c in df_export.columns]]
        
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df_export.to_excel(writer, index=False, sheet_name="機器台帳マスター")
        excel_data = output.getvalue()
        
        file_name = "機器台帳マスター.xlsx"
        
        if mode == "2. 全自動（データベース保存）":
            encoded = base64.b64encode(excel_data).decode("utf-8")
            api_url = f"https://api.github.com/repos/{repo}/contents/ledger/{urllib.parse.quote(file_name)}"
            
            sha = None
            try:
                req_check = urllib.request.Request(api_url)
                req_check.add_header("Authorization", f"token {token}")
                with urllib.request.urlopen(req_check) as res:
                    sha = json.loads(res.read().decode("utf-8"))["sha"]
            except: pass
            
            payload = {"message": "Update Master Ledger", "content": encoded, "branch": "main"}
            if sha: payload["sha"] = sha
            
            req = urllib.request.Request(api_url, data=json.dumps(payload).encode("utf-8"), method="PUT")
            req.add_header("Authorization", f"token {token}")
            req.add_header("Content-Type", "application/json")
            urllib.request.urlopen(req)
            
        elif mode == "3. 社内共有フォルダへ自動保存":
            target_dir = Path(local_path)
            target_dir.mkdir(parents=True, exist_ok=True)
            out_path = target_dir / file_name
            with open(out_path, "wb") as f:
                f.write(excel_data)
    except Exception as e:
        print(f"Excelマスター台帳の保存エラー: {e}")

# ==========================================
# --- メインアプリ ---
# ==========================================
def main():
    qp = st.query_params
    is_redirect_mode = "id" in qp
    
    if is_redirect_mode:
        st.set_page_config(page_title="機器情報ページ", layout="centered")
        st.markdown("<style>#MainMenu {visibility: hidden;} footer {visibility: hidden;}</style>", unsafe_allow_html=True)

        target_id = qp["id"]
        if DB_CSV.exists():
            try:
                df = pd.read_csv(DB_CSV)
                match = df[df["ID"].astype(str) == str(target_id)]
                if not match.empty:
                    target_url = match.iloc[-1]["URL"]
                    link_html = f"""
                    <div style="text-align: center; margin-top: 60px;">
                        <p style="font-size: 20px; font-weight: bold; color: #333;">✅ 機器情報ページの準備ができました</p>
                        <a href="{target_url}" style="display:inline-block; margin-top:15px; padding:20px 40px; background:#28a745; color:white; font-size:22px; font-weight:bold; text-decoration:none; border-radius:8px; box-shadow:0 4px 6px rgba(0,0,0,0.2);">
                            📱 機器情報ページを開く
                        </a>
                    </div>
                    """
                    st.markdown(link_html, unsafe_allow_html=True)
                else: st.error("エラー: 管理番号が見つかりません。")
            except: st.error("データベース読み込みエラー")
        else: st.error("データベースが見つかりません。")
        return

    st.set_page_config(page_title="機器情報ページ ＆ QR管理システム", layout="wide", initial_sidebar_state="expanded")
    
    st.markdown("""
    <style>
    .stButton button { width: 100%; border-radius: 5px; }
    .block-container { padding-top: 1.5rem !important; }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] { gap: 0.3rem !important; }
    [data-testid="stSidebar"] button { padding: 0 !important; height: 32px !important; min-height: 32px !important; display: flex; align-items: center; justify-content: center; }
    #MainMenu {visibility: hidden;} footer {visibility: hidden;}
    </style>
    """, unsafe_allow_html=True)

    db_columns = ["ID", "Name", "Power", "URL", "Updated", "memo", "img_exterior", "img_outlet", "img_label", "img_loto1", "img_loto2", "extra_images"]
    if not DB_CSV.exists():
        pd.DataFrame(columns=db_columns).to_csv(DB_CSV, index=False)
    else:
        df_init = pd.read_csv(DB_CSV)
        needs_save = False
        for col in db_columns:
            if col not in df_init.columns:
                df_init[col] = ""
                needs_save = True
        if needs_save: df_init.to_csv(DB_CSV, index=False)

    if "input_did" not in st.session_state: st.session_state.input_did = ""
    if "input_name" not in st.session_state: st.session_state.input_name = ""
    if "input_power" not in st.session_state: st.session_state.input_power = None
    if "input_memo" not in st.session_state: st.session_state.input_memo = ""
    if "is_related_loto" not in st.session_state: st.session_state.is_related_loto = False
    
    if "form_reset_key" not in st.session_state: st.session_state.form_reset_key = 0
    if "extra_images_count" not in st.session_state: st.session_state.extra_images_count = 0
    rk = st.session_state.form_reset_key

    if "current_db_sel" not in st.session_state: st.session_state.current_db_sel = "✨ 新規登録 (クリア)"

    # ==========================================
    # --- 【重要】削除ボタン用の魔法のコールバック関数 ---
    # ==========================================
    def delete_db_item_callback(did_to_del):
        try:
            d_csv = pd.read_csv(DB_CSV)
            d_csv = d_csv[d_csv["ID"].astype(str) != str(did_to_del)]
            d_csv.to_csv(DB_CSV, index=False)
        except Exception:
            pass
            
        st.session_state.input_did = ""
        st.session_state.input_name = ""
        st.session_state.input_power = None
        st.session_state.input_memo = ""
        st.session_state.is_related_loto = False
        st.session_state.existing_imgs = {}
        st.session_state.existing_ex_imgs = []
        st.session_state.extra_images_count = 0
        st.session_state.current_db_sel = "✨ 新規登録 (クリア)"
        
        # 掟破りにならない安全なリセット方法
        if "db_select_widget" in st.session_state:
            st.session_state.db_select_widget = "✨ 新規登録 (クリア)"
            
        st.session_state.form_reset_key += 1
        st.session_state["scroll_to_top"] = True
        st.session_state.delete_success_msg = True


    st.sidebar.header("🗄️ 登録済み機器データベース")
    if DB_CSV.exists():
        df = pd.read_csv(DB_CSV)
        if not df.empty:
            options = ["✨ 新規登録 (クリア)"] + (df["ID"].astype(str) + " : " + df["Name"]).tolist()
            
            c_sel = st.session_state.current_db_sel
            sel_idx = options.index(c_sel) if c_sel in options else 0
            
            selected_edit = st.sidebar.selectbox("編集・確認する機器を選択:", options, index=sel_idx, key="db_select_widget")
            
            if selected_edit != st.session_state.current_db_sel:
                st.session_state.current_db_sel = selected_edit
                if selected_edit == "✨ 新規登録 (クリア)":
                    st.session_state.input_did = ""
                    st.session_state.input_name = ""
                    st.session_state.input_power = None
                    st.session_state.input_memo = ""
                    st.session_state.is_related_loto = False
                    st.session_state.existing_imgs = {}
                    st.session_state.existing_ex_imgs = []
                    st.session_state.extra_images_count = 0
                else:
                    did_str = selected_edit.split(" : ")[0]
                    match = df[df["ID"].astype(str) == did_str]
                    if not match.empty:
                        row = match.iloc[-1]
                        st.session_state.input_did = str(row["ID"])
                        st.session_state.input_name = str(row["Name"])
                        p_val = str(row.get("Power", "")) if pd.notna(row.get("Power")) else None
                        st.session_state.input_power = p_val if p_val in ["100V", "200V"] else None
                        st.session_state.input_memo = str(row.get("memo", "")) if pd.notna(row.get("memo")) else ""
                        st.session_state.is_related_loto = False
                        
                        st.session_state.existing_imgs = {
                            "ext": str(row.get("img_exterior", "")), "out": str(row.get("img_outlet", "")),
                            "lab": str(row.get("img_label", "")), "lo1": str(row.get("img_loto1", "")), "lo2": str(row.get("img_loto2", ""))
                        }
                        ex_str = str(row.get("extra_images", "[]"))
                        if pd.isna(row.get("extra_images")): ex_str = "[]"
                        try: st.session_state.existing_ex_imgs = json.loads(ex_str)
                        except: st.session_state.existing_ex_imgs = []
                st.session_state.form_reset_key += 1
                st.rerun()

            if st.session_state.current_db_sel != "✨ 新規登録 (クリア)":
                st.sidebar.info("💡 過去の画像とデータが呼び出されました。そのまま再発行や、一部の画像の差し替えが可能です。")
                did_val = st.session_state.current_db_sel.split(" : ")[0]
                
                # 【エラー完全解消】削除ボタンもコールバック関数の魔法に対応！
                st.sidebar.button("🗑️ この機器をデータベースから削除", on_click=delete_db_item_callback, args=(did_val,))
                
            if st.session_state.get("delete_success_msg"):
                st.sidebar.success("✅ 削除しました！")
                st.session_state.delete_success_msg = False

    st.sidebar.markdown("---")
    st.sidebar.header("⚙️ システム詳細設定")
    save_mode = st.sidebar.radio("保存モードを選択:", ["1. 手動ダウンロードのみ", "2. 全自動（データベース保存）", "3. 社内共有フォルダへ自動保存"], index=1)
    
    github_repo = ""; github_token = ""; local_path = ""
    if save_mode == "2. 全自動（データベース保存）":
        github_repo = st.sidebar.text_input("データベース領域名", value="equipment-portal/qr-manager")
        github_token = st.sidebar.text_input("システム接続キー (トークン)", value=st.secrets.get("github_token", ""), type="password")
    elif save_mode == "3. 社内共有フォルダへ自動保存":
        local_path = st.sidebar.text_input("共有フォルダのパス", value=r"C:\Equipment_Manuals")

    st.sidebar.markdown("---")
    st.sidebar.markdown("**⏬ 手動保存オプション**")
    include_equip_name = st.sidebar.checkbox(
        "プレビュー画像の保存名に「機器名称」を含める", 
        value=True, 
        help="プレビュー確認後に、手動で画像をPCへダウンロードする際のファイル名に適用されます（例: 2699_金型反転機.jpg）"
    )
    
    st.markdown("<div id='top_anchor'></div>", unsafe_allow_html=True)
    st.title("📱 機器情報ページ ＆ QR管理システム")
    
    if st.session_state.get("scroll_to_top"):
        import streamlit.components.v1 as components
        components.html("<script>var t=window.parent.document.getElementById('top_anchor');if(t){t.scrollIntoView(true);}else{window.parent.scrollTo(0,0);}</script>", height=0)
        st.session_state["scroll_to_top"] = False
        
    col1, col2 = st.columns(2)
    
    with col1:
        st.header("1. 基本情報入力")
        did = st.text_input("管理番号 (例: 2699)", key="input_did")
        name = st.text_input("機器名称 (例: 5t金型反転機)", key="input_name")
        power = st.selectbox("使用電源", ["100V", "200V"], index=None, key="input_power")
        
        st.markdown("---")
        st.header("📝 メモ・備考欄")
        memo = st.text_area("現場へ伝える補足情報", height=150, key="input_memo")

    def render_image_ui(label, key_suffix, existing_path):
        st.markdown(f"**{label}**")
        has_existing = pd.notna(existing_path) and str(existing_path).strip() != "" and str(existing_path) != "nan"
        
        del_flag = False
        if has_existing:
            try:
                if str(existing_path).startswith("http"): st.image(existing_path, width=180, caption="現在保存されている画像")
                elif Path(existing_path).exists(): st.image(str(existing_path), width=180, caption="現在保存されている画像")
                del_flag = st.checkbox(f"🗑️ この画像を削除する", key=f"del_{key_suffix}_{rk}")
            except:
                st.warning("※保存先の画像が見つかりません")
                has_existing = False
                
        new_file = st.file_uploader("新しい画像で上書きする" if has_existing else "画像をアップロード", type=["png", "jpg", "jpeg"], key=f"up_{key_suffix}_{rk}")
        st.markdown("<hr style='margin:10px 0;'>", unsafe_allow_html=True)
        return new_file, del_flag, existing_path if has_existing else ""

    with col2:
        st.header("2. 画像の指定・管理")
        imgs = st.session_state.get("existing_imgs", {})
        
        f_ext, d_ext, e_ext = render_image_ui("機器外観", "ext", imgs.get("ext", ""))
        f_out, d_out, e_out = render_image_ui("コンセント位置", "out", imgs.get("out", ""))
        f_lab, d_lab, e_lab = render_image_ui("資産管理ラベル", "lab", imgs.get("lab", ""))
        
        st.markdown("---")
        is_related_loto = st.checkbox("関連機器、付帯設備のLOTO手順書を登録する", key="is_related_loto")
        f_lo1, d_lo1, e_lo1 = render_image_ui("LOTO手順書（1ページ目）", "lo1", imgs.get("lo1", ""))
        f_lo2, d_lo2, e_lo2 = render_image_ui("LOTO手順書（2ページ目）", "lo2", imgs.get("lo2", ""))
        
        st.markdown("---")
        st.subheader("➕ 追加情報の画像")
        
        ex_imgs_data_preview = [] 
        ex_imgs_to_save = [] 
        
        existing_ex = st.session_state.get("existing_ex_imgs", [])
        
        for i, ex_dict in enumerate(existing_ex):
            st.markdown(f"**既存の追加画像 {i+1}**")
            e_path = ex_dict.get("url", "")
            e_title = ex_dict.get("title", f"追加画像 {i+1}")
            
            if str(e_path).startswith("http"): st.image(e_path, width=180)
            elif Path(e_path).exists(): st.image(str(e_path), width=180)
            
            del_ex = st.checkbox(f"🗑️ この追加画像を削除", key=f"del_ex_{rk}_{i}")
            new_title = st.text_input(f"タイトル変更", value=e_title, key=f"edit_ex_t_{rk}_{i}")
            new_f = st.file_uploader(f"画像を差し替え", type=["png", "jpg", "jpeg"], key=f"edit_ex_f_{rk}_{i}")
            st.markdown("<hr style='margin:10px 0;'>", unsafe_allow_html=True)
            
            if not del_ex:
                final_f = new_f if new_f else e_path
                ex_imgs_data_preview.append((final_f, new_title))
                ex_imgs_to_save.append({"type": "existing", "file": new_f, "url": e_path, "title": new_title, "index": i})
        
        for i in range(st.session_state.extra_images_count):
            st.markdown(f"**新規の追加画像 {len(existing_ex) + i + 1}**")
            et = st.text_input(f"タイトル", key=f"new_ex_title_{rk}_{i}")
            ef = st.file_uploader(f"画像", type=["png", "jpg", "jpeg"], key=f"new_ex_img_{rk}_{i}")
            st.markdown("<hr style='margin:10px 0;'>", unsafe_allow_html=True)
            if ef:
                t = et if et else f"追加画像 {len(existing_ex) + i + 1}"
                ex_imgs_data_preview.append((ef, t))
                ex_imgs_to_save.append({"type": "new", "file": ef, "title": t, "index": i})

        if st.button("➕ 追加枠を増やす"):
            st.session_state.extra_images_count += 1
            st.rerun()

    def get_input_for_manual(file_obj, del_flag, existing_path):
        if del_flag: return None
        if file_obj: return file_obj
        if existing_path: return existing_path
        return None

    # --- プレビュー機能 ---
    st.markdown("---")
    st.header("3. 機器情報ページ プレビュー確認")
    if st.button("🔍 プレビューを作成", type="secondary"):
        if did and name and power:
            with st.spinner("プレビュー作成中..."):
                m_data = {
                    "id": did, "name": name, "power": power, "memo": memo, "is_related_loto": is_related_loto,
                    "img_exterior": get_input_for_manual(f_ext, d_ext, e_ext),
                    "img_outlet": get_input_for_manual(f_out, d_out, e_out),
                    "img_label": get_input_for_manual(f_lab, d_lab, e_lab),
                    "img_loto1": get_input_for_manual(f_lo1, d_lo1, e_lo1),
                    "img_loto2": get_input_for_manual(f_lo2, d_lo2, e_lo2)
                }
                manual_path = MANUAL_DIR / f"preview_{rk}.jpg"
                create_manual_image_extended(m_data, ex_imgs_data_preview, manual_path)
                if manual_path.exists():
                    st.success("プレビュー成功！")
                    with open(manual_path, "rb") as f: img_b64 = base64.b64encode(f.read()).decode("utf-8")
                    st.components.v1.html(f'<div style="height:500px; overflow-y:scroll; border:2px solid #ddd;"><img src="data:image/jpeg;base64,{img_b64}" width="100%"></div>', height=520)
                    
                    s_id = safe_filename(did)
                    dl_file_name = f"{s_id}_{safe_filename(name)}.jpg" if include_equip_name else f"{s_id}.jpg"
                    with open(manual_path, "rb") as img_file:
                        st.download_button(label="📥 完成したプレビュー画像を手動でPCに保存", data=img_file, file_name=dl_file_name, mime="image/jpeg")
        else:
            st.error("管理番号、機器名称、使用電源は必須です。")

    # --- 登録・発行機能 ---
    st.markdown("---")
    st.header("4. データ登録 ＆ 印刷用ラベル発行")
    
    if save_mode == "1. 手動ダウンロードのみ":
        long_url = st.text_input("保管先等のURLを貼り付け")
        if st.button("🖨️ 手動設定でラベルを発行", type="primary"):
            if long_url and did and name and power:
                qr_path = QR_DIR / f"{safe_filename(did)}_qr.png"
                img_qr = make_optimized_qr(long_url)
                img_qr.save(qr_path)
                
                df = pd.read_csv(DB_CSV)
                df = df[df["ID"].astype(str) != str(did)]
                new_data = {"ID": did, "Name": name, "Power": power, "URL": long_url, "Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "memo": memo}
                df = pd.concat([df, pd.DataFrame([new_data])], ignore_index=True)
                df.to_csv(DB_CSV, index=False)
                
                label_img = create_label_image({"name": name, "power": power, "img_qr": img_qr})
                add_label_to_history(name, label_img)
                st.image(label_img, caption="印刷用ラベル", width=300)

    elif save_mode in ["2. 全自動（データベース保存）", "3. 社内共有フォルダへ自動保存"]:
        if st.button("🖨️ 【全自動】画像を保存し、台帳へ登録してラベルを発行する", type="primary"):
            if did and name and power:
                with st.spinner("🔄 画像の圧縮とデータベース保存を実行中..."):
                    try:
                        def process_save(f_obj, d_flag, e_path, suffix):
                            if d_flag: return ""
                            if f_obj: return save_image_to_storage(f_obj, did, suffix, save_mode, github_repo, github_token, local_path)
                            return e_path
                            
                        fin_ext = process_save(f_ext, d_ext, e_ext, "ext")
                        fin_out = process_save(f_out, d_out, e_out, "out")
                        fin_lab = process_save(f_lab, d_lab, e_lab, "lab")
                        fin_lo1 = process_save(f_lo1, d_lo1, e_lo1, "lo1")
                        fin_lo2 = process_save(f_lo2, d_lo2, e_lo2, "lo2")

                        final_extra_images_db = []
                        for item in ex_imgs_to_save:
                            if item["type"] == "existing":
                                if item["file"]: 
                                    saved_url = save_image_to_storage(item["file"], did, f"ex_{item['index']}", save_mode, github_repo, github_token, local_path)
                                    if saved_url: final_extra_images_db.append({"title": item["title"], "url": saved_url})
                                else: 
                                    final_extra_images_db.append({"title": item["title"], "url": item["url"]})
                            elif item["type"] == "new":
                                if item["file"]: 
                                    saved_url = save_image_to_storage(item["file"], did, f"ex_new_{item['index']}", save_mode, github_repo, github_token, local_path)
                                    if saved_url: final_extra_images_db.append({"title": item["title"], "url": saved_url})

                        m_data = {
                            "id": did, "name": name, "power": power, "memo": memo, "is_related_loto": is_related_loto,
                            "img_exterior": fin_ext if fin_ext else None,
                            "img_outlet": fin_out if fin_out else None,
                            "img_label": fin_lab if fin_lab else None,
                            "img_loto1": fin_lo1 if fin_lo1 else None,
                            "img_loto2": fin_lo2 if fin_lo2 else None
                        }
                        
                        s_id = safe_filename(did)
                        # タイムスタンプを外し、常に「管理番号.jpg」で上書き保存する！
                        file_name_manual = f"{s_id}.jpg"
                        manual_path = MANUAL_DIR / file_name_manual
                        
                        create_manual_image_extended(m_data, ex_imgs_data_preview, manual_path)

                        final_manual_url = ""
                        if save_mode == "2. 全自動（データベース保存）":
                            with open(manual_path, "rb") as f:
                                payload = {"message": f"Upload Manual {file_name_manual}", "content": base64.b64encode(f.read()).decode("utf-8"), "branch": "main"}
                                req = urllib.request.Request(f"https://api.github.com/repos/{github_repo}/contents/manuals/{urllib.parse.quote(file_name_manual)}", data=json.dumps(payload).encode("utf-8"), method="PUT")
                                req.add_header("Authorization", f"token {github_token}"); req.add_header("Content-Type", "application/json")
                                with urllib.request.urlopen(req) as res:
                                    final_manual_url = json.loads(res.read().decode("utf-8"))["content"]["html_url"].replace("https://github.com/", "https://cdn.jsdelivr.net/gh/").replace("/blob/", "@")
                                
                                # 【追加】更新時に、クラウドサーバー(CDN)の古い画像を強制的に吹き飛ばす！
                                try:
                                    urllib.request.urlopen(f"https://purge.jsdelivr.net/gh/{github_repo}@main/manuals/{file_name_manual}")
                                except: pass
                        
                        elif save_mode == "3. 社内共有フォルダへ自動保存":
                            target_dir = Path(local_path) / "manuals"
                            target_dir.mkdir(parents=True, exist_ok=True)
                            out_manual = target_dir / file_name_manual
                            shutil.copy(manual_path, out_manual)
                            final_manual_url = str(out_manual).replace("\\", "/")

                        qr_path = QR_DIR / f"{s_id}_qr.png"
                        img_qr = make_optimized_qr(final_manual_url)
                        img_qr.save(qr_path)
                        
                        label_img = create_label_image({"name": name, "power": power, "img_qr": img_qr})
                        add_label_to_history(name, label_img)

                        df = pd.read_csv(DB_CSV)
                        df = df[df["ID"].astype(str) != str(did)]
                        new_row = {
                            "ID": did, "Name": name, "Power": power, "URL": final_manual_url, "Updated": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                            "memo": memo, "img_exterior": fin_ext, "img_outlet": fin_out, "img_label": fin_lab, "img_loto1": fin_lo1, "img_loto2": fin_lo2,
                            "extra_images": json.dumps(final_extra_images_db, ensure_ascii=False) 
                        }
                        df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
                        df.to_csv(DB_CSV, index=False)

                        update_master_ledger_excel(df, save_mode, github_repo, github_token, local_path)

                        st.success(f"✅ 登録完了！ マニュアルURL: {final_manual_url}")
                        st.image(label_img, caption="印刷用ラベル", width=300)
                        
                    except Exception as e:
                        st.error(f"エラーが発生しました: {str(e)}")

    # ==========================================
    # --- 【最強進化】環境まるごとバックアップ保存＆復元 ---
    # ==========================================
    
    def reset_form_callback():
        st.session_state.input_did = ""
        st.session_state.input_name = ""
        st.session_state.input_power = None
        st.session_state.input_memo = ""
        st.session_state.is_related_loto = False
        st.session_state.existing_imgs = {}
        st.session_state.existing_ex_imgs = []
        st.session_state.extra_images_count = 0
        st.session_state.current_db_sel = "✨ 新規登録 (クリア)"
        if "db_select_widget" in st.session_state:
            st.session_state.db_select_widget = "✨ 新規登録 (クリア)"
        st.session_state.form_reset_key += 1
        st.session_state["scroll_to_top"] = True

    def restore_backup_callback():
        current_rk = st.session_state.form_reset_key
        uploaded_file = st.session_state.get(f"backup_up_{current_rk}")
        if uploaded_file is not None:
            try:
                loaded_data = json.loads(uploaded_file.getvalue().decode("utf-8"))
                
                form_data = loaded_data.get("form", loaded_data) 
                workspace_data = loaded_data.get("workspace", {})

                if workspace_data:
                    if "devices_csv" in workspace_data and workspace_data["devices_csv"]:
                        with open(DB_CSV, "w", encoding="utf-8") as f:
                            f.write(workspace_data["devices_csv"])
                    
                    label_imgs = workspace_data.get("label_images", {})
                    for img_name, b64_str in label_imgs.items():
                        try:
                            with open(TEMP_LABEL_DIR / img_name, "wb") as f:
                                f.write(base64.b64decode(b64_str))
                        except: pass

                    label_hist = workspace_data.get("label_history", [])
                    with open(LABEL_HISTORY_FILE, "w", encoding="utf-8") as f:
                        json.dump(label_hist, f, ensure_ascii=False, indent=2)
                    
                    rebuild_excel()

                st.session_state.input_did = form_data.get("did", "")
                st.session_state.input_name = form_data.get("name", "")
                p_val = form_data.get("power", "")
                st.session_state.input_power = p_val if p_val in ["100V", "200V"] else None
                st.session_state.input_memo = form_data.get("memo", "")
                st.session_state.is_related_loto = form_data.get("is_related_loto", False)
                
                def decode_img(img_dict, prefix):
                    if not img_dict: return ""
                    if img_dict["type"] == "path": return img_dict["data"]
                    if img_dict["type"] == "base64":
                        try:
                            b_data = base64.b64decode(img_dict["data"])
                            temp_path = DRAFT_IMG_DIR / f"restored_{prefix}_{datetime.now().strftime('%H%M%S%f')}.jpg"
                            with open(temp_path, "wb") as f: f.write(b_data)
                            return str(temp_path)
                        except: return ""
                    return ""

                restored_imgs = {}
                d_imgs = form_data.get("existing_imgs", {})
                restored_imgs["ext"] = decode_img(d_imgs.get("ext"), "ext")
                restored_imgs["out"] = decode_img(d_imgs.get("out"), "out")
                restored_imgs["lab"] = decode_img(d_imgs.get("lab"), "lab")
                restored_imgs["lo1"] = decode_img(d_imgs.get("lo1"), "lo1")
                restored_imgs["lo2"] = decode_img(d_imgs.get("lo2"), "lo2")
                st.session_state.existing_imgs = restored_imgs
                
                restored_ex = []
                for i, ex in enumerate(form_data.get("existing_ex_imgs", [])):
                    path = decode_img(ex.get("img_data"), f"ex_{i}")
                    restored_ex.append({"title": ex.get("title", ""), "url": path})
                st.session_state.existing_ex_imgs = restored_ex
                
                st.session_state.extra_images_count = 0
                st.session_state.current_db_sel = "✨ 新規登録 (クリア)"
                
                if "db_select_widget" in st.session_state:
                    st.session_state.db_select_widget = "✨ 新規登録 (クリア)"
                    
                st.session_state.form_reset_key += 1
                st.session_state["scroll_to_top"] = True
                
                st.session_state.restore_success = True
            except Exception as e:
                st.session_state.backup_error_msg = f"バックアップの読み込みに失敗しました: {e}"

    st.markdown("---")
    st.header("5. ワークスペース（作業状態のバックアップ） ＆ 次の作業")
    
    if st.session_state.get("restore_success"):
        st.success("✅ バックアップから「入力状態」「データベース」「Excel台帳」をすべて復元しました！")
        st.session_state.restore_success = False

    col_a, col_b = st.columns(2)
    
    with col_a:
        st.subheader("📝 作業状態をまるごとPCに保存")
        st.info("入力中の文字・画像だけでなく、左側の「データベース」や右側の「Excel台帳」も含めて、今の環境をそのままファイルとしてPCにダウンロードします。")
        
        form_draft = {
            "did": did, "name": name, "power": power, "memo": memo, "is_related_loto": is_related_loto,
            "existing_imgs": {}, "existing_ex_imgs": [], "extra_images_count": st.session_state.extra_images_count
        }
        
        def encode_img_fixed(f_obj, e_path):
            if f_obj:
                try:
                    comp_bytes = compress_image(f_obj) 
                    if comp_bytes:
                        return {"type": "base64", "data": base64.b64encode(comp_bytes).decode("utf-8")}
                except: pass
            elif e_path:
                return {"type": "path", "data": str(e_path)}
            return None

        form_draft["existing_imgs"]["ext"] = encode_img_fixed(f_ext, e_ext)
        form_draft["existing_imgs"]["out"] = encode_img_fixed(f_out, e_out)
        form_draft["existing_imgs"]["lab"] = encode_img_fixed(f_lab, e_lab)
        form_draft["existing_imgs"]["lo1"] = encode_img_fixed(f_lo1, e_lo1)
        form_draft["existing_imgs"]["lo2"] = encode_img_fixed(f_lo2, e_lo2)

        ex_imgs = []
        for item in ex_imgs_to_save:
            enc = encode_img_fixed(item.get("file"), item.get("url", ""))
            ex_imgs.append({"title": item.get("title", ""), "img_data": enc})
        form_draft["existing_ex_imgs"] = ex_imgs
        
        csv_data_str = ""
        if DB_CSV.exists():
            with open(DB_CSV, "r", encoding="utf-8") as f: csv_data_str = f.read()

        label_hist_data = []
        if LABEL_HISTORY_FILE.exists():
            try:
                with open(LABEL_HISTORY_FILE, "r", encoding="utf-8") as f: label_hist_data = json.load(f)
            except: pass

        label_images_b64 = {}
        for item in label_hist_data:
            img_name = item.get("img_filename")
            img_p = TEMP_LABEL_DIR / img_name
            if img_p.exists():
                with open(img_p, "rb") as f: label_images_b64[img_name] = base64.b64encode(f.read()).decode("utf-8")

        backup_data = {
            "form": form_draft,
            "workspace": {
                "devices_csv": csv_data_str,
                "label_history": label_hist_data,
                "label_images": label_images_b64
            }
        }
        
        backup_json_str = json.dumps(backup_data, ensure_ascii=False)
        dl_filename = f"backup_{safe_filename(did) if did else 'workspace'}_{datetime.now().strftime('%Y%m%d%H%M')}.json"
        
        st.download_button(
            label="💾 現在の状態を【ワークスペース保存(.json)】としてPCに保存",
            data=backup_json_str,
            file_name=dl_filename,
            mime="application/json",
            use_container_width=True
        )
        
        st.markdown("---")
        st.subheader("📂 PCに保存したバックアップを復元")
        uploaded_backup = st.file_uploader("保存したファイル(.json)を選択", type=["json"], key=f"backup_up_{rk}")
        if uploaded_backup:
            st.button("🔄 このバックアップ環境を復元する", type="primary", use_container_width=True, on_click=restore_backup_callback)
            
        if "backup_error_msg" in st.session_state:
            st.error(st.session_state.backup_error_msg)
            del st.session_state.backup_error_msg

    with col_b:
        st.subheader("🔄 登録の完了・リセット")
        st.info("💡 **印刷用台帳の状況は常に自動保存されています！** ブラウザを閉じても、明日はそのまま続きからラベル印刷が可能です。")
        
        st.button("🔄 次の機器を入力する (クリアして上へ戻る)", type="primary", use_container_width=True, on_click=reset_form_callback)

    # --- サイドバー：Excel台帳状況 ---
    st.sidebar.markdown("---")
    st.sidebar.subheader("🖨️ 印刷用Excel台帳の状況")
    h_list = []
    if LABEL_HISTORY_FILE.exists():
        try:
            with open(LABEL_HISTORY_FILE, "r", encoding="utf-8") as f: h_list = json.load(f)
        except: pass
    
    c_len = len(h_list)
    if c_len == 0: st.sidebar.info("🈳 現在、台帳は白紙です。")
    else:
        st.sidebar.success(f"✅ 合計 {c_len} 枚のラベルを配置済み")
        cols = ((c_len - 1) // 13) + 1
        grid_html = "<div style='background:#f0f2f6;padding:10px;border-radius:5px;font-size:13px;line-height:1.2;text-align:left;'>"
        for r in range(13):
            line = ""
            for c in range(cols):
                idx = c * 13 + r
                if idx < c_len:
                    num_icon = chr(9311 + idx + 1) if idx < 20 else f"({idx+1})"
                    line += f"<span style='display:inline-block;width:26px;text-align:center;font-weight:bold;color:#d4af37;'>{num_icon}</span>"
                else: line += "<span style='display:inline-block;width:26px;text-align:center;color:#ccc;'>⬜</span>"
            grid_html += line + "<br>"
        st.sidebar.markdown(grid_html + "</div>", unsafe_allow_html=True)
        
        for i, obj in enumerate(h_list):
            cb1, cb2 = st.sidebar.columns([5, 1])
            icon = chr(9311 + i + 1) if i < 20 else f"({i+1})"
            cb1.markdown(f"<div style='display: flex; align-items: center; height: 32px; font-size: 15px;'>{icon} {obj['name']}</div>", unsafe_allow_html=True)
            if cb2.button("❌", key=f"d_itm_{i}"): delete_label_from_history(i); st.rerun()
    
    if EXCEL_LABEL_PATH.exists():
        with open(EXCEL_LABEL_PATH, "rb") as f: st.sidebar.download_button("📥 最新のExcelをダウンロード", f, "labels.xlsx")
        if st.sidebar.button("🗑️ 台帳をリセット"): clear_history(); st.rerun()

if __name__ == "__main__":
    main()

